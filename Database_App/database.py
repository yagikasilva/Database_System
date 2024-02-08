from customtkinter import *
from tkinter import *
from openpyxl import load_workbook

app = CTk()
app.geometry("700x400")
app.title("Database System")
app.iconbitmap('databaseimage.ico')


frame_right = CTkScrollableFrame(app,fg_color="#242424",width=380,height=390)
frame_right.place(x=300,y=0)

tabview=CTkTabview(app,fg_color="white",height=400)
tabview.place(x=0,y=0)

tabview.add("Students")
tabview.add("Personal Details")


# Create a label with the image
heading = CTkLabel(tabview.tab("Students"), text="Database System",text_color="black",font=("Arial",20,"bold"))
heading.place(x=70, y=10)

i_Name = CTkEntry (tabview.tab("Students"),width=250,placeholder_text="Your Name")
i_Name.place(x=28,y=90)

i_Grade = CTkEntry (tabview.tab("Students"),width=250,placeholder_text="Grade")
i_Grade.place(x=28,y=130)




heading = CTkLabel(tabview.tab("Personal Details"), text="Database System\nPersonal Deails",text_color="black",font=("Arial",20,"bold"))
heading.place(x=70, y=10)
#tab 2 entryes
i_Address = CTkComboBox (tabview.tab("Personal Details"),values=["Male","Female"],width=250)
i_Address.place(x=28,y=90)

i_Age= CTkEntry (tabview.tab("Personal Details"),width=250,placeholder_text="Your Age")
i_Age.place(x=28,y=130)

i_Tele = CTkEntry (tabview.tab("Personal Details"),width=250,placeholder_text="Your Telephone Number")
i_Tele.place(x=28,y=170)
# Create a button


def close():
    tabview.tab("Personal Details").quit()

def function(input_Name,input_Grade,):
        
    wb = load_workbook("password.xlsx")
    ws = wb["Sheet4"]
    
    # Append the value to the worksheet
    ws.append([input_Name,input_Grade])

    # Display the values in the right frame
    x = 2
    p = ws["A" + str(x)].value
    q = ws["B" + str(x)].value
    
    while p is not None:
        column_heading1=CTkLabel(frame_right, text="Name", text_color="white", font=("Helvetica",20,"bold"))
        column_heading1.grid(column=1, row=1,padx=4)
        column_heading2=CTkLabel(frame_right, text="Age", text_color="white",font=("Helvetica",20,"bold"))
        column_heading2.grid(column=2, row=1,padx=4)
        
        result_label = CTkLabel(frame_right, text=f"{p}", text_color="white",font=("Tahoma",13))
        result_label.grid(column=1, row=(x))
        
        result_label2 = CTkLabel(frame_right, text=f"{q}", text_color="white",font=("Tahoma",13))
        result_label2.grid(column=2, row=(x))  # Adjust the y-coordinate for spacing
        
        x += 1
        p = ws["A" + str(x)].value
        q = ws["B" + str(x)].value
        
    
    # Save the changes to the workbook
    wb.save("password.xlsx")
    

def function2(input_Name,input_Grade,input_Address,input_Age,input_Tele):
        
    wb = load_workbook("password.xlsx")
    ws = wb["Sheet4"]
    
    # Get the value from the entry widget
    #i_Name_value = entry_widget.get()

    # Append the value to the worksheet
    ws.append([input_Name,input_Grade,input_Address,input_Age,input_Tele])

    # Display the values in the right frame
    x = 2
    p = ws["A" + str(x)].value
    q = ws["B" + str(x)].value
    r = ws["C" + str(x)].value
    s = ws["D" + str(x)].value
    t = ws["E" + str(x)].value
    
    while p is not None:
        column_heading1=CTkLabel(frame_right, text="Name", text_color="white", font=("Helvetica",20,"bold"))
        column_heading1.grid(column=1, row=1,padx=4)
        column_heading2=CTkLabel(frame_right, text="Grade", text_color="white",font=("Helvetica",20,"bold"))
        column_heading2.grid(column=2, row=1,padx=4)
        column_heading2=CTkLabel(frame_right, text="Address", text_color="white",font=("Helvetica",20,"bold"))
        column_heading2.grid(column=3, row=1,padx=4)
        column_heading2=CTkLabel(frame_right, text="Age", text_color="white",font=("Helvetica",20,"bold"))
        column_heading2.grid(column=4, row=1,padx=4)
        column_heading2=CTkLabel(frame_right, text="Tele No.", text_color="white",font=("Helvetica",20,"bold"))
        column_heading2.grid(column=5, row=1,padx=4)
        
        result_label = CTkLabel(frame_right, text=f"{p}", text_color="white",font=("Tahoma",13))
        result_label.grid(column=1, row=(x),padx=4)
        
        result_label2 = CTkLabel(frame_right, text=f"{q}", text_color="white",font=("Tahoma",13))
        result_label2.grid(column=2, row=(x),padx=4) 
        
        result_label3 = CTkLabel(frame_right, text=f"{r}", text_color="white",font=("Tahoma",13))
        result_label3.grid(column=3, row=(x),padx=4) # Adjust the y-coordinate for spacing

        result_label4 = CTkLabel(frame_right, text=f"{s}", text_color="white",font=("Tahoma",13))
        result_label4.grid(column=4, row=(x),padx=4) 

        result_label5 = CTkLabel(frame_right, text=f"{t}", text_color="white",font=("Tahoma",13))
        result_label5.grid(column=5, row=(x),padx=4) 
                
        x += 1
        p = ws["A" + str(x)].value
        q = ws["B" + str(x)].value
        r = ws["C" + str(x)].value
        s = ws["D" + str(x)].value
        t = ws["E" + str(x)].value
        
    
    # Save the changes to the workbook
    wb.save("password.xlsx")
    

    
btn = CTkButton(tabview.tab("Students"), text="Upload",command=lambda:function(i_Name.get(),i_Grade.get()))
btn.place(x=70, y=300)
btn2 = CTkButton(tabview.tab("Personal Details"), text="Upload",command=lambda:function2(i_Name.get(),i_Grade.get(),i_Address.get(),i_Age.get(),i_Tele.get()))
btn2.place(x=70, y=300)

close1 = CTkButton(tabview.tab("Personal Details"), text="X",command=close,width=1,height=20,fg_color="red",corner_radius=100)
close1.place(x=250, y=0)


app.mainloop()


